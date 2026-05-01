import streamlit as st
import openpyxl
from openai import OpenAI
from dotenv import load_dotenv
import os
import io

try:
    load_dotenv(encoding="utf-8")
except UnicodeDecodeError:
    # Fallback for ANSI/GBK encoded .env files on Windows.
    load_dotenv(encoding="gbk")

st.set_page_config(page_title="Automotive Requirements Analyzer", layout="wide")
st.title("🚗 Automotive Requirements Intelligent Analyzer")
st.markdown("Upload a requirements Excel file. AI will automatically classify, score quality, and suggest improvements.")

with st.sidebar:
    st.header("⚙️ Settings")
    api_key = st.text_input("API Key", type="password", placeholder="Enter your API Key")
    model = st.selectbox("Select Model", [
        "deepseek-ai/DeepSeek-V3",
        "google/gemini-2.0-flash-001"
    ])
    st.markdown("---")
    st.markdown("**How to Use**")
    st.markdown("1. Enter your API Key")
    st.markdown("2. Upload Excel file (Column A = requirements)")
    st.markdown("3. Click Start Analysis")
    st.markdown("4. Download results")

uploaded_file = st.file_uploader("Upload Requirements Excel File", type=["xlsx"])

if uploaded_file and api_key:
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    requirements = []
    for row in range(1, ws.max_row + 1):
        val = ws[f"A{row}"].value
        if val:
            requirements.append(str(val))

    st.success(f"✅ Successfully loaded {len(requirements)} requirements")

    with st.expander("📋 View Requirements List"):
        for i, r in enumerate(requirements):
            st.write(f"{i+1}. {r}")

    if st.button("🚀 Start Analysis", type="primary"):
        client = OpenAI(api_key=api_key, base_url="https://api.siliconflow.cn/v1")

        with st.spinner("AI is categorizing requirement types..."):
            all_reqs = "\n".join([f"{i+1}. {r}" for i, r in enumerate(requirements)])
            response = client.chat.completions.create(
                model=model,
                temperature=0.1,
                messages=[{"role": "user", "content": f"""You are an automotive requirements engineering expert.
Based on the following requirements, derive the most appropriate requirement types from the data itself.
Do not use any predefined categories.
Provide a name and short definition for each type.

Requirements:
{all_reqs}

Reply format:
Type1: xxx | Definition: xxx
Type2: xxx | Definition: xxx"""}]
            )
            categories = response.choices[0].message.content

        st.subheader("📊 AI-Derived Requirement Types")
        st.text(categories)
        st.markdown("---")

        st.subheader("📝 Analysis Results")
        results = []
        progress = st.progress(0)

        for i, req in enumerate(requirements):
            with st.spinner(f"Analyzing requirement {i+1}/{len(requirements)}..."):
                response = client.chat.completions.create(
                    model=model,
                    temperature=0.1,
                    messages=[{"role": "user", "content": f"""You are an automotive requirements engineering expert.
Analyze the following requirement based on the type definitions below.

Available types:
{categories}

Requirement: {req}

Reply format (strictly follow this format):
Type: xxx
Score: integer from 1 to 5
Suggestion: xxx"""}]
                )
                result = response.choices[0].message.content

                req_type, score, suggestion = "", "", ""
                for line in result.strip().split("\n"):
                    if line.startswith("Type:"):
                        req_type = line.replace("Type:", "").strip()
                    elif line.startswith("Score:"):
                        score = line.replace("Score:", "").strip()
                    elif line.startswith("Suggestion:"):
                        suggestion = line.replace("Suggestion:", "").strip()

                results.append({
                    "Requirement": req,
                    "AI Type": req_type,
                    "Quality Score": score,
                    "Improvement Suggestion": suggestion
                })

                with st.container():
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        st.markdown(f"**{req}**")
                        st.caption(f"Suggestion: {suggestion}")
                    with col2:
                        st.metric("Type", req_type)
                        st.metric("Score", score)
                    st.markdown("---")

                progress.progress((i + 1) / len(requirements))

        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.append(["Requirement", "AI Type", "Quality Score", "Improvement Suggestion"])
        for r in results:
            ws_out.append([r["Requirement"], r["AI Type"], r["Quality Score"], r["Improvement Suggestion"]])

        buffer = io.BytesIO()
        wb_out.save(buffer)
        buffer.seek(0)

        st.success("🎉 Analysis Complete!")
        st.download_button(
            label="📥 Download Analysis Results (Excel)",
            data=buffer,
            file_name="requirements_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

elif uploaded_file and not api_key:
    st.warning("⚠️ Please enter your API Key in the sidebar")
elif not uploaded_file:
    st.info("👆 Please upload an Excel file to begin")